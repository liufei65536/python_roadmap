import openpyxl
from openpyxl import Workbook, load_workbook

# 文档：https://openpyxl.readthedocs.io/en/stable/tutorial.html
# 中文文档（较长时间未更新）https://openpyxl-chinese-docs.readthedocs.io/zh_CN/latest/tutorial.html


def read_save_access():
    # 读取 .xlsx
    wb = load_workbook('./data/scores.xlsx')
    # 获取当前work sheet
    ws = wb.active
    # 访问work sheet的单元
    print(ws['B1'].value)
    ws['B1'].value = "语文2"

    # 获得当前所有的sheet
    print(wb.sheetnames)
    # 切换sheet
    ws = wb['Sheet2']
    # 新增sheet
    wb.create_sheet("add sheet")

    # 保存wb到文件
    # wb.save("./data/scores2.xlsx")

def insert_pic():
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'You should see three logos below'
    img = Image('py.png')
    ws.add_image(img, 'A1')
    wb.save('logo.xlsx')

if __name__ == '__main__':
    insert_pic()
